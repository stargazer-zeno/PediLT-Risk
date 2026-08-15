from __future__ import annotations

import json
import logging
from io import BytesIO
from pathlib import Path
from typing import Any, Iterable

LOGGER = logging.getLogger(__name__)

# Canonical sections of a standard patient JSON.
SECTION_BASE = "基础信息"
SECTION_FOLLOWUP = "时序随访基础信息"
SECTION_LAB = "时序检验指标 (纯数值序列)"
SECTION_MED = "时序用药记录 (纯数值序列)"

# Column-name prefixes used to route flat tabular columns into nested sections.
# Both English and Chinese prefixes are accepted, separator ":" or "："
FOLLOWUP_PREFIXES = ("fu:", "fu：", "随访:", "随访：")
LAB_PREFIXES = ("lab:", "lab：", "检验:", "检验：")
MED_PREFIXES = ("med:", "med：", "用药:", "用药：")

# Multiple timesteps inside one tabular cell may be separated by these.
SERIES_SEPARATORS = (";", "；", "|")


class StandardDataAdapterError(ValueError):
    pass


class StandardDataAdapter:
    """Convert heterogeneous uploads into standard patient JSON objects.

    Supported inputs:

    * ``.json``   -- a single patient object, a list of objects, or an object
                     wrapping ``{"patients": [...]}`` / ``{"data": [...]}``.
    * ``.jsonl``  -- one patient JSON object per line.
    * ``.csv`` / ``.xlsx`` -- one patient per row. A ``patient_json`` column may
                     carry a full JSON string; otherwise columns are routed to
                     sections by name (see prefixes above). Time-series cells may
                     contain several timesteps separated by ``;`` ``；`` or ``|``.
    """

    def from_bytes(self, content: bytes, filename: str) -> list[dict[str, Any]]:
        suffix = Path(filename).suffix.lower()
        if suffix == ".json":
            return self._from_json_text(content.decode("utf-8-sig"))
        if suffix == ".jsonl":
            return self._from_jsonl_text(content.decode("utf-8-sig"))
        if suffix == ".csv":
            return self._from_table(self._read_csv(content))
        if suffix in {".xlsx", ".xls"}:
            return self._from_table(self._read_excel(content))
        raise StandardDataAdapterError(
            f"Unsupported file type {suffix!r}. Use .json, .jsonl, .csv or .xlsx."
        )

    # ------------------------------------------------------------------ JSON
    def from_json_obj(self, obj: Any) -> list[dict[str, Any]]:
        if isinstance(obj, dict):
            for wrapper in ("patients", "data", "records", "items"):
                if wrapper in obj and isinstance(obj[wrapper], list):
                    return [self._coerce_obj(o) for o in obj[wrapper]]
            return [self._coerce_obj(obj)]
        if isinstance(obj, list):
            if not obj:
                raise StandardDataAdapterError("Input list is empty.")
            return [self._coerce_obj(o) for o in obj]
        raise StandardDataAdapterError("JSON input must be an object or a list.")

    def _from_json_text(self, text: str) -> list[dict[str, Any]]:
        try:
            obj = json.loads(text)
        except json.JSONDecodeError as exc:
            raise StandardDataAdapterError(f"Invalid JSON: {exc}") from exc
        return self.from_json_obj(obj)

    def _from_jsonl_text(self, text: str) -> list[dict[str, Any]]:
        records: list[dict[str, Any]] = []
        for lineno, line in enumerate(text.splitlines(), start=1):
            line = line.strip()
            if not line:
                continue
            try:
                records.append(self._coerce_obj(json.loads(line)))
            except json.JSONDecodeError as exc:
                raise StandardDataAdapterError(
                    f"Invalid JSON on line {lineno}: {exc}"
                ) from exc
        if not records:
            raise StandardDataAdapterError("JSONL input contained no records.")
        return records

    def _coerce_obj(self, obj: Any) -> dict[str, Any]:
        if not isinstance(obj, dict):
            raise StandardDataAdapterError("Each patient record must be a JSON object.")
        return self._ensure_sections(obj)

    # ----------------------------------------------------------------- Tables
    def _read_csv(self, content: bytes) -> list[dict[str, Any]]:
        import csv

        text = content.decode("utf-8-sig")
        reader = csv.DictReader(text.splitlines())
        return [dict(row) for row in reader]

    def _read_excel(self, content: bytes) -> list[dict[str, Any]]:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:  # pragma: no cover
            raise StandardDataAdapterError(
                "openpyxl is required to read .xlsx files."
            ) from exc

        wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        try:
            header = [str(h) if h is not None else "" for h in next(rows)]
        except StopIteration:
            return []
        records: list[dict[str, Any]] = []
        for row in rows:
            if row is None or all(c is None for c in row):
                continue
            record = {}
            for key, value in zip(header, row):
                if key:
                    record[key] = value
            records.append(record)
        wb.close()
        return records

    def _from_table(self, rows: Iterable[dict[str, Any]]) -> list[dict[str, Any]]:
        records = list(rows)
        if not records:
            raise StandardDataAdapterError("Tabular input contained no rows.")
        patients = [self._row_to_patient(row) for row in records]
        return patients

    def _row_to_patient(self, row: dict[str, Any]) -> dict[str, Any]:
        # Full-JSON escape hatch.
        for col in ("patient_json", "json", "standard_json"):
            if col in row and _nonempty(row[col]):
                try:
                    return self._ensure_sections(json.loads(str(row[col])))
                except json.JSONDecodeError as exc:
                    raise StandardDataAdapterError(
                        f"Invalid JSON in column {col!r}: {exc}"
                    ) from exc

        base: dict[str, Any] = {}
        followup: list[str] = []
        labs: list[str] = []
        meds: list[str] = []
        patient_id: Any = None

        for raw_key, value in row.items():
            if raw_key is None:
                continue
            key = str(raw_key).strip()
            if not key or not _nonempty(value):
                continue

            if key.lower() in {"id", "patient_id", "节点id", "编号"}:
                patient_id = value
                continue

            series = _matched_prefix(key, FOLLOWUP_PREFIXES)
            if series is not None:
                followup.append(_series_field(series, value))
                continue
            series = _matched_prefix(key, LAB_PREFIXES)
            if series is not None:
                labs.append(_series_field(series, value))
                continue
            series = _matched_prefix(key, MED_PREFIXES)
            if series is not None:
                meds.append(_series_field(series, value))
                continue

            # Unprefixed columns are treated as static base-info fields.
            base[key] = value

        patient: dict[str, Any] = {SECTION_BASE: base}
        if patient_id is not None:
            patient["id"] = patient_id
        patient[SECTION_FOLLOWUP] = followup
        patient[SECTION_LAB] = labs
        patient[SECTION_MED] = meds
        return patient

    # --------------------------------------------------------------- helpers
    def _ensure_sections(self, obj: dict[str, Any]) -> dict[str, Any]:
        obj.setdefault(SECTION_BASE, {})
        obj.setdefault(SECTION_FOLLOWUP, [])
        obj.setdefault(SECTION_LAB, [])
        obj.setdefault(SECTION_MED, [])
        return obj


def _nonempty(value: Any) -> bool:
    if value is None:
        return False
    return str(value).strip() != ""


def _matched_prefix(key: str, prefixes: tuple[str, ...]) -> str | None:
    lowered = key.lower()
    for prefix in prefixes:
        if lowered.startswith(prefix.lower()):
            return key[len(prefix):].strip()
    return None


def _series_field(name: str, value: Any) -> str:
    """Render a series cell as the canonical ``"NAME: v1, v2。"`` string."""
    text = str(value).strip()
    for sep in SERIES_SEPARATORS:
        text = text.replace(sep, ",")
    parts = [p.strip() for p in text.split(",") if p.strip()]
    joined = ", ".join(parts) if parts else text
    return f"{name}: {joined}。"
